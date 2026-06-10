from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

SUMMARY_HEADERS = (
    "Cost Code",
    "Category",
    "Cost Items",
    "Budget",
    "Approved Contract",
    "Forecast Variations",
    "Approved Variations",
    "Forecast Final Cost",
    "Budget Variance",
    "Claimed to Date",
    "This Month",
    "Remaining",
)
INVOICE_HEADERS = (
    "Invoice Date",
    "Company",
    "PO Number",
    "Invoice Number",
    "Invoice Description",
    "Cost Item",
    "Amount",
    "Billing Month",
    "Paid?",
)
VARIATION_HEADERS = (
    "Date Submitted",
    "Cost Item",
    "Variation To",
    "Status",
    "Amount",
    "Date Approved",
    "Approved Amount",
)
CATEGORY_ORDER = (
    "Fees and charges",
    "Consultants",
    "Construction",
    "PC allowances",
    "Contingency / allowances",
)
SUMMARY_MONEY_COLUMNS = set(range(4, 13))
INVOICE_MONEY_COLUMNS = {7}
VARIATION_MONEY_COLUMNS = {5, 7}
PREVIEW_MAX_ROWS = 160


@dataclass(frozen=True, slots=True)
class CostPlanLine:
    cost_code: str
    category: str
    cost_item: str
    budget: float | None
    approved_contract: float | None
    status: str
    basis: str


@dataclass(frozen=True, slots=True)
class CostPlanWorkbook:
    content: bytes
    filename: str
    row_count: int
    cost_item_lookup_count: int
    warnings: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class WorkbookSheetPreview:
    name: str
    column_count: int
    rows: list[list[str]]
    styles: list[list[dict[str, Any]]]


@dataclass(frozen=True, slots=True)
class WorkbookPreview:
    sheets: list[WorkbookSheetPreview]
    warnings: list[str]


def build_cost_plan_workbook(
    *,
    project_title: str,
    markdown: str,
    version: int,
    generated_at: datetime | None = None,
) -> CostPlanWorkbook:
    generated_at = generated_at or datetime.now(UTC)
    items, warnings = parse_cost_breakdown(markdown)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    invoices = wb.create_sheet("Invoices")
    variations = wb.create_sheet("Variations")
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    _build_summary_sheet(summary, project_title, generated_at, items)
    _build_invoices_sheet(invoices, project_title)
    _build_variations_sheet(variations, project_title)
    _add_defined_names(wb, max(len(items) + 1, 2))
    _verify_workbook(wb)

    buffer = BytesIO()
    wb.save(buffer)
    return CostPlanWorkbook(
        content=buffer.getvalue(),
        filename=f"Cost_Plan_v{version:02d}.draft.xlsx",
        row_count=len(items),
        cost_item_lookup_count=len(items),
        warnings=tuple(warnings),
    )


def parse_cost_breakdown(markdown: str) -> tuple[list[CostPlanLine], list[str]]:
    section = _markdown_section(markdown, "Cost breakdown by category")
    if not section:
        return [], ["Cost breakdown by category section was not found."]

    table = [line.strip() for line in section if _is_markdown_table_line(line)]
    if len(table) < 2:
        return [], ["Cost breakdown by category table was not found."]

    headers = [_normalise_header(cell) for cell in _split_markdown_row(table[0])]
    header_index = {header: index for index, header in enumerate(headers)}
    required = ("cost code", "category", "cost items", "budget")
    missing = [header for header in required if header not in header_index]
    if missing:
        return [], [f"Cost breakdown table is missing columns: {', '.join(missing)}."]

    rows: list[CostPlanLine] = []
    warnings: list[str] = []
    last_category = ""
    for raw in table[1:]:
        cells = _split_markdown_row(raw)
        if _is_separator_row(cells):
            continue

        cost_item = _clean_markdown_cell(_cell(cells, header_index["cost items"]))
        category = _clean_markdown_cell(_cell(cells, header_index["category"]))
        cost_code = _clean_markdown_cell(_cell(cells, header_index["cost code"]))
        if _is_total_row(cost_code, category, cost_item):
            continue
        if not cost_item:
            continue

        if category:
            last_category = category
        category = _normalise_category(category or last_category)
        budget = _parse_money(_cell(cells, header_index["budget"]))
        status = _clean_markdown_cell(_cell(cells, header_index.get("status", -1)))
        basis = _clean_markdown_cell(_cell(cells, header_index.get("basis", -1)))

        approved_contract = None
        approved_index = header_index.get("approved contract")
        if approved_index is not None:
            approved_contract = _parse_money(_cell(cells, approved_index))
        elif _status_marks_approved(status):
            approved_contract = budget

        rows.append(
            CostPlanLine(
                cost_code=cost_code or str(len(rows) + 1),
                category=category,
                cost_item=cost_item,
                budget=budget,
                approved_contract=approved_contract,
                status=status,
                basis=basis,
            )
        )

    if not rows:
        warnings.append("Cost breakdown table did not contain any workbook cost item rows.")
    return rows, warnings


def workbook_preview_from_bytes(content: bytes) -> WorkbookPreview:
    workbook = load_workbook(BytesIO(content), data_only=False)
    rollup = _summary_rollup_values(workbook)
    warnings: list[str] = []
    sheets: list[WorkbookSheetPreview] = []
    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            continue
        sheets.append(_sheet_preview(worksheet, rollup))

    expected = ["Summary", "Invoices", "Variations"]
    if workbook.sheetnames[:3] != expected:
        warnings.append("Workbook tabs differ from the expected Summary, Invoices, Variations layout.")

    return WorkbookPreview(sheets=sheets, warnings=warnings)


def _build_summary_sheet(
    worksheet: Worksheet,
    project_title: str,
    generated_at: datetime,
    items: list[CostPlanLine],
) -> None:
    fills = _fills()
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"
    worksheet.merge_cells("A1:L1")
    worksheet.merge_cells("A2:L2")
    worksheet["A1"] = f"Project Cost Plan - {project_title}"
    worksheet["A2"] = "All figures exclude GST"
    worksheet["J3"] = "Selected month"
    worksheet["K3"] = _month_start(generated_at.date())
    worksheet["K3"].number_format = "mmm-yy"

    for cell in worksheet[1]:
        cell.fill = fills["title"]
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
    worksheet["A2"].fill = fills["subtitle"]
    worksheet["A2"].font = Font(italic=True, color="44546A")
    worksheet["J3"].font = Font(bold=True, color="44546A")
    worksheet["K3"].fill = fills["control"]
    worksheet["K3"].font = Font(bold=True)
    worksheet["K3"].alignment = Alignment(horizontal="center")

    for column, header in enumerate(SUMMARY_HEADERS, start=1):
        cell = worksheet.cell(row=4, column=column, value=header)
        cell.fill = fills["header"]
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    grouped = _group_cost_items(items)
    row_number = 5
    subtotal_rows: list[int] = []
    for category, category_items in grouped:
        if not category_items:
            continue
        start_row = row_number
        for item in category_items:
            _write_summary_item_row(worksheet, row_number, item)
            row_number += 1
        subtotal_rows.append(row_number)
        _write_summary_total_row(worksheet, row_number, category, start_row, row_number - 1)
        row_number += 2

    _write_summary_grand_total_row(worksheet, row_number, subtotal_rows)
    _write_summary_lookup_columns(worksheet, items, generated_at)
    _add_summary_validations(worksheet)
    _style_summary_sheet(worksheet, row_number)


def _build_invoices_sheet(worksheet: Worksheet, project_title: str) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"
    worksheet.merge_cells("A1:I1")
    worksheet.merge_cells("A2:I2")
    worksheet["A1"] = "INVOICES REGISTER"
    worksheet["A2"] = project_title
    _style_title_rows(worksheet, "I")
    for column, header in enumerate(INVOICE_HEADERS, start=1):
        worksheet.cell(row=4, column=column, value=header)
    _style_register_headers(worksheet, len(INVOICE_HEADERS))
    _set_widths(
        worksheet,
        {
            "A": 14,
            "B": 24,
            "C": 15,
            "D": 18,
            "E": 34,
            "F": 30,
            "G": 14,
            "H": 14,
            "I": 10,
        },
    )
    for row in range(5, 501):
        worksheet.cell(row=row, column=1).number_format = "dd-mmm-yy"
        worksheet.cell(row=row, column=7).number_format = "$#,##0"
        worksheet.cell(row=row, column=8).number_format = "mmm-yy"
    _add_list_validation(worksheet, "F5:F500", "CostItemLookup", "Invalid Cost Item")
    _add_list_validation(worksheet, "I5:I500", '"Yes,No"', "Invalid Paid value")


def _build_variations_sheet(worksheet: Worksheet, project_title: str) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"
    worksheet.merge_cells("A1:G1")
    worksheet.merge_cells("A2:G2")
    worksheet["A1"] = "VARIATIONS REGISTER"
    worksheet["A2"] = project_title
    _style_title_rows(worksheet, "G")
    for column, header in enumerate(VARIATION_HEADERS, start=1):
        worksheet.cell(row=4, column=column, value=header)
    _style_register_headers(worksheet, len(VARIATION_HEADERS))
    _set_widths(
        worksheet,
        {
            "A": 16,
            "B": 30,
            "C": 24,
            "D": 16,
            "E": 14,
            "F": 16,
            "G": 16,
        },
    )
    for row in range(5, 501):
        worksheet.cell(row=row, column=1).number_format = "dd-mmm-yy"
        worksheet.cell(row=row, column=5).number_format = "$#,##0"
        worksheet.cell(row=row, column=6).number_format = "dd-mmm-yy"
        worksheet.cell(row=row, column=7).number_format = "$#,##0"
    _add_list_validation(worksheet, "B5:B500", "CostItemLookup", "Invalid Cost Item")
    _add_list_validation(
        worksheet,
        "D5:D500",
        '"Forecast,Submitted,Withdrawn,Rejected,Approved"',
        "Invalid Status",
    )


def _write_summary_item_row(worksheet: Worksheet, row: int, item: CostPlanLine) -> None:
    values: dict[int, Any] = {
        1: item.cost_code,
        2: item.category,
        3: item.cost_item,
        4: item.budget,
        5: item.approved_contract,
        6: f'=SUMIFS(Variations!$E$5:$E$500,Variations!$B$5:$B$500,C{row},Variations!$D$5:$D$500,"<>Approved")',
        7: f"=SUMIF(Variations!$B$5:$B$500,C{row},Variations!$G$5:$G$500)",
        8: f"=SUM(E{row}:G{row})",
        9: f"=D{row}-H{row}",
        10: f'=SUMIFS(Invoices!$G$5:$G$500,Invoices!$F$5:$F$500,C{row},Invoices!$H$5:$H$500,"<="&EOMONTH(K$3,0))',
        11: f'=SUMIFS(Invoices!$G$5:$G$500,Invoices!$F$5:$F$500,C{row},Invoices!$H$5:$H$500,">="&EOMONTH(K$3,-1)+1,Invoices!$H$5:$H$500,"<="&EOMONTH(K$3,0))',
        12: f"=D{row}-J{row}",
    }
    for column, value in values.items():
        worksheet.cell(row=row, column=column, value=value)


def _write_summary_total_row(
    worksheet: Worksheet,
    row: int,
    category: str,
    start_row: int,
    end_row: int,
) -> None:
    worksheet.cell(row=row, column=2, value=category)
    worksheet.cell(row=row, column=3, value="Subtotal")
    for column in range(4, 13):
        letter = get_column_letter(column)
        worksheet.cell(row=row, column=column, value=f"=SUM({letter}{start_row}:{letter}{end_row})")


def _write_summary_grand_total_row(
    worksheet: Worksheet,
    row: int,
    subtotal_rows: list[int],
) -> None:
    worksheet.cell(row=row, column=2, value="Grand total")
    worksheet.cell(row=row, column=3, value="All cost items")
    for column in range(4, 13):
        letter = get_column_letter(column)
        if subtotal_rows:
            refs = ",".join(f"{letter}{subtotal_row}" for subtotal_row in subtotal_rows)
            worksheet.cell(row=row, column=column, value=f"=SUM({refs})")
        else:
            worksheet.cell(row=row, column=column, value=0)


def _write_summary_lookup_columns(
    worksheet: Worksheet,
    items: list[CostPlanLine],
    generated_at: datetime,
) -> None:
    worksheet["M1"] = "Cost Item Lookup"
    worksheet["N1"] = "Invoice Billing Months"
    for index, item in enumerate(items, start=2):
        worksheet.cell(row=index, column=13, value=item.cost_item)

    anchor = _month_start(generated_at.date())
    for index, month_offset in enumerate(range(-12, 48), start=2):
        cell = worksheet.cell(row=index, column=14, value=_add_months(anchor, month_offset))
        cell.number_format = "mmm-yy"

    worksheet.column_dimensions["M"].hidden = True
    worksheet.column_dimensions["N"].hidden = True


def _add_defined_names(workbook: Workbook, last_lookup_row: int) -> None:
    workbook.defined_names.add(
        DefinedName(
            "CostItemLookup",
            attr_text=f"'Summary'!$M$2:$M${last_lookup_row}",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "InvoiceBillingMonths",
            attr_text="'Summary'!$N$2:$N$61",
        )
    )


def _add_summary_validations(worksheet: Worksheet) -> None:
    _add_list_validation(worksheet, "K3", "InvoiceBillingMonths", "Invalid Billing Month")


def _add_list_validation(
    worksheet: Worksheet,
    cell_range: str,
    formula: str,
    error_title: str,
) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.showDropDown = False
    validation.showErrorMessage = False
    validation.errorTitle = error_title
    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def _style_summary_sheet(worksheet: Worksheet, last_row: int) -> None:
    fills = _fills()
    thin_border = _thin_border()
    widths = {
        "A": 10,
        "B": 21,
        "C": 34,
        "D": 14,
        "E": 17,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 16,
        "J": 16,
        "K": 14,
        "L": 14,
    }
    _set_widths(worksheet, widths)
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[4].height = 34

    for row in worksheet.iter_rows(min_row=4, max_row=last_row, min_col=1, max_col=12):
        is_total = _is_summary_total_cells(row)
        is_grand = _cell_text(row[1].value).lower() == "grand total"
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=cell.column in {2, 3},
                horizontal="right" if cell.column in SUMMARY_MONEY_COLUMNS else "left",
            )
            if cell.column in SUMMARY_MONEY_COLUMNS and cell.row >= 5:
                cell.number_format = "$#,##0"
            if is_total:
                cell.fill = fills["subtotal"]
                cell.font = Font(bold=True)
            if is_grand:
                cell.fill = fills["grand"]
                cell.font = Font(bold=True)

    for row in worksheet.iter_rows(min_row=1, max_row=61, min_col=13, max_col=14):
        for cell in row:
            cell.fill = fills["lookup"]


def _style_title_rows(worksheet: Worksheet, final_column: str) -> None:
    fills = _fills()
    for cell in worksheet[f"A1:{final_column}1"][0]:
        cell.fill = fills["title"]
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
    worksheet["A2"].fill = fills["subtitle"]
    worksheet["A2"].font = Font(italic=True, color="44546A")
    worksheet["A2"].alignment = Alignment(horizontal="center")


def _style_register_headers(worksheet: Worksheet, column_count: int) -> None:
    fills = _fills()
    border = _thin_border()
    for row in worksheet.iter_rows(min_row=4, max_row=500, min_col=1, max_col=column_count):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=cell.column not in {7},
                horizontal="right" if cell.column in {5, 7} else "left",
            )
    for cell in worksheet[4]:
        cell.fill = fills["header"]
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[4].height = 32


def _fills() -> dict[str, PatternFill]:
    return {
        "title": PatternFill("solid", fgColor="1F4E78"),
        "subtitle": PatternFill("solid", fgColor="D9EAF7"),
        "header": PatternFill("solid", fgColor="44546A"),
        "control": PatternFill("solid", fgColor="FFF2CC"),
        "subtotal": PatternFill("solid", fgColor="E2F0D9"),
        "grand": PatternFill("solid", fgColor="BDD7EE"),
        "lookup": PatternFill("solid", fgColor="F2F2F2"),
    }


def _thin_border() -> Border:
    side = Side(style="thin", color="D9E2F3")
    return Border(left=side, right=side, top=side, bottom=side)


def _set_widths(worksheet: Worksheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _verify_workbook(workbook: Workbook) -> None:
    issues: list[str] = []
    if workbook.sheetnames != ["Summary", "Invoices", "Variations"]:
        issues.append("Workbook tabs must be exactly Summary, Invoices, Variations.")
    for name in ("CostItemLookup", "InvoiceBillingMonths"):
        if name not in workbook.defined_names:
            issues.append(f"Workbook is missing named range {name}.")

    expected_validations = {
        "Summary": ("K3", "InvoiceBillingMonths"),
        "Invoices": ("F5:F500", "CostItemLookup"),
        "Variations": ("B5:B500", "CostItemLookup"),
    }
    for sheet_name, (cell_range, formula) in expected_validations.items():
        if not _has_data_validation(workbook[sheet_name], cell_range, formula):
            issues.append(f"{sheet_name} is missing data validation for {cell_range}.")

    summary = workbook["Summary"]
    for row in range(5, summary.max_row + 1):
        formula = summary.cell(row=row, column=6).value
        if isinstance(formula, str) and "#REF!" in formula:
            issues.append(f"Summary row {row} contains a broken formula reference.")

    if issues:
        raise ValueError(" ".join(issues))


def _has_data_validation(worksheet: Worksheet, cell_range: str, formula: str) -> bool:
    for validation in worksheet.data_validations.dataValidation:
        if cell_range in str(validation.sqref) and validation.formula1 == formula:
            return True
    return False


def _sheet_preview(
    worksheet: Worksheet,
    rollup: dict[tuple[str, int, int], Any],
) -> WorkbookSheetPreview:
    visible_columns = [
        column
        for column in range(1, worksheet.max_column + 1)
        if not worksheet.column_dimensions[get_column_letter(column)].hidden
    ]
    last_row = min(_last_visible_row(worksheet, visible_columns, rollup), PREVIEW_MAX_ROWS)
    last_column_index = _last_visible_column(worksheet, visible_columns, last_row, rollup)
    visible_columns = [column for column in visible_columns if column <= last_column_index]

    rows: list[list[str]] = []
    styles: list[list[dict[str, Any]]] = []
    for row in range(1, last_row + 1):
        preview_row: list[str] = []
        style_row: list[dict[str, Any]] = []
        for column in visible_columns:
            cell = worksheet.cell(row=row, column=column)
            value = rollup.get((worksheet.title, row, column), cell.value)
            preview_row.append(_display_cell_value(worksheet.title, row, column, value))
            style_row.append(_cell_style(cell))
        rows.append(preview_row)
        styles.append(style_row)

    return WorkbookSheetPreview(
        name=worksheet.title,
        column_count=len(visible_columns),
        rows=rows,
        styles=styles,
    )


def _summary_rollup_values(workbook: Workbook) -> dict[tuple[str, int, int], Any]:
    if not {"Summary", "Invoices", "Variations"}.issubset(set(workbook.sheetnames)):
        return {}

    summary = workbook["Summary"]
    invoices = workbook["Invoices"]
    variations = workbook["Variations"]
    selected_month = _month_start(_coerce_date(summary["K3"].value) or date.today())

    invoice_amounts = _invoice_amounts_by_cost_item(invoices)
    variation_amounts = _variation_amounts_by_cost_item(variations)
    rollup: dict[tuple[str, int, int], Any] = {}
    item_rows: list[int] = []
    current_group_rows: list[int] = []

    header_row = _summary_header_row(summary)
    if header_row is None:
        return rollup

    for row in range(header_row + 1, summary.max_row + 1):
        cost_item = _cell_text(summary.cell(row=row, column=3).value)
        row_label = f"{summary.cell(row=row, column=2).value or ''} {cost_item}".lower()
        if not row_label.strip():
            continue
        if "subtotal" in row_label:
            for column in range(4, 13):
                rollup[("Summary", row, column)] = sum(
                    _summary_number(summary, rollup, item_row, column)
                    for item_row in current_group_rows
                )
            current_group_rows = []
            continue
        if "grand total" in row_label:
            for column in range(4, 13):
                rollup[("Summary", row, column)] = sum(
                    _summary_number(summary, rollup, item_row, column)
                    for item_row in item_rows
                )
            continue
        if not cost_item:
            continue

        budget = _number(summary.cell(row=row, column=4).value) or 0
        approved_contract = _number(summary.cell(row=row, column=5).value) or 0
        forecast_variations = variation_amounts[cost_item]["forecast"]
        approved_variations = variation_amounts[cost_item]["approved"]
        claimed_to_date = sum(
            amount
            for month, amount in invoice_amounts[cost_item]
            if month <= selected_month
        )
        this_month = sum(
            amount
            for month, amount in invoice_amounts[cost_item]
            if month == selected_month
        )
        forecast_final = approved_contract + forecast_variations + approved_variations

        rollup[("Summary", row, 6)] = forecast_variations
        rollup[("Summary", row, 7)] = approved_variations
        rollup[("Summary", row, 8)] = forecast_final
        rollup[("Summary", row, 9)] = budget - forecast_final
        rollup[("Summary", row, 10)] = claimed_to_date
        rollup[("Summary", row, 11)] = this_month
        rollup[("Summary", row, 12)] = budget - claimed_to_date
        item_rows.append(row)
        current_group_rows.append(row)

    return rollup


def _invoice_amounts_by_cost_item(worksheet: Worksheet) -> dict[str, list[tuple[date, float]]]:
    amounts: dict[str, list[tuple[date, float]]] = defaultdict(list)
    for row in range(5, worksheet.max_row + 1):
        cost_item = _cell_text(worksheet.cell(row=row, column=6).value)
        amount = _number(worksheet.cell(row=row, column=7).value)
        billing_month = _coerce_date(worksheet.cell(row=row, column=8).value)
        if cost_item and amount is not None and billing_month is not None:
            amounts[cost_item].append((_month_start(billing_month), amount))
    return amounts


def _variation_amounts_by_cost_item(worksheet: Worksheet) -> dict[str, dict[str, float]]:
    amounts: dict[str, dict[str, float]] = defaultdict(lambda: {"forecast": 0.0, "approved": 0.0})
    for row in range(5, worksheet.max_row + 1):
        cost_item = _cell_text(worksheet.cell(row=row, column=2).value)
        if not cost_item:
            continue
        status = _cell_text(worksheet.cell(row=row, column=4).value).lower()
        amount = _number(worksheet.cell(row=row, column=5).value) or 0
        approved_amount = _number(worksheet.cell(row=row, column=7).value) or 0
        if status == "approved":
            amounts[cost_item]["approved"] += approved_amount
        else:
            amounts[cost_item]["forecast"] += amount
            amounts[cost_item]["approved"] += approved_amount
    return amounts


def _summary_header_row(worksheet: Worksheet) -> int | None:
    for row in range(1, min(worksheet.max_row, 20) + 1):
        if (
            _cell_text(worksheet.cell(row=row, column=1).value) == "Cost Code"
            and _cell_text(worksheet.cell(row=row, column=3).value) == "Cost Items"
        ):
            return row
    return None


def _summary_number(
    worksheet: Worksheet,
    rollup: dict[tuple[str, int, int], Any],
    row: int,
    column: int,
) -> float:
    value = rollup.get(("Summary", row, column), worksheet.cell(row=row, column=column).value)
    return _number(value) or 0


def _last_visible_row(
    worksheet: Worksheet,
    columns: list[int],
    rollup: dict[tuple[str, int, int], Any],
) -> int:
    last_row = 1
    for row in range(1, worksheet.max_row + 1):
        if any(
            _cell_text(rollup.get((worksheet.title, row, column), worksheet.cell(row=row, column=column).value))
            for column in columns
        ):
            last_row = row
    return last_row


def _last_visible_column(
    worksheet: Worksheet,
    columns: list[int],
    last_row: int,
    rollup: dict[tuple[str, int, int], Any],
) -> int:
    last_column = columns[0] if columns else 1
    for column in columns:
        if any(
            _cell_text(rollup.get((worksheet.title, row, column), worksheet.cell(row=row, column=column).value))
            for row in range(1, last_row + 1)
        ):
            last_column = column
    return last_column


def _cell_style(cell: Cell) -> dict[str, Any]:
    fill_color = _argb_to_hex(cell.fill.fgColor.rgb)
    return {
        "fill_color": fill_color,
        "bold": bool(cell.font and cell.font.bold),
    }


def _display_cell_value(sheet_name: str, row: int, column: int, value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        if value.startswith("="):
            return ""
        return value
    if isinstance(value, datetime):
        return _display_date(sheet_name, row, column, value.date())
    if isinstance(value, date):
        return _display_date(sheet_name, row, column, value)
    if isinstance(value, int | float):
        if _is_money_preview_cell(sheet_name, row, column):
            return _format_money(value)
        return f"{value:g}"
    return str(value)


def _display_date(sheet_name: str, row: int, column: int, value: date) -> str:
    if sheet_name == "Summary" and row == 3 and column == 11:
        return value.strftime("%b-%y")
    if sheet_name == "Invoices" and column == 8:
        return value.strftime("%b-%y")
    return value.strftime("%d %b %Y")


def _is_money_preview_cell(sheet_name: str, row: int, column: int) -> bool:
    if row < 5:
        return False
    if sheet_name == "Summary":
        return column in SUMMARY_MONEY_COLUMNS
    if sheet_name == "Invoices":
        return column in INVOICE_MONEY_COLUMNS
    if sheet_name == "Variations":
        return column in VARIATION_MONEY_COLUMNS
    return False


def _format_money(value: float) -> str:
    if abs(value) < 0.005:
        return ""
    sign = "-" if value < 0 else ""
    return f"{sign}${abs(value):,.0f}"


def _markdown_section(markdown: str, heading: str) -> list[str]:
    target = heading.strip().lower()
    lines = markdown.splitlines()
    section: list[str] = []
    inside = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("## "):
            current = stripped[3:].strip().lower()
            if inside and current != target:
                break
            inside = current == target
            continue
        if inside:
            section.append(line)
    return section


def _is_markdown_table_line(line: str) -> bool:
    stripped = line.strip()
    return stripped.startswith("|") and stripped.endswith("|")


def _split_markdown_row(row: str) -> list[str]:
    stripped = row.strip().strip("|")
    return [cell.strip() for cell in stripped.split("|")]


def _cell(cells: list[str], index: int) -> str:
    if index < 0 or index >= len(cells):
        return ""
    return cells[index]


def _normalise_header(value: str) -> str:
    value = _clean_markdown_cell(value).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value).strip()
    aliases = {
        "cost item": "cost items",
        "item": "cost items",
        "amount": "budget",
        "approved": "approved contract",
    }
    return aliases.get(value, value)


def _clean_markdown_cell(value: str) -> str:
    cleaned = re.sub(r"(\*\*|__|`)", "", value)
    cleaned = cleaned.replace("&amp;", "&")
    return " ".join(cleaned.split())


def _is_separator_row(cells: list[str]) -> bool:
    return all(not cell.replace("-", "").replace(":", "").strip() for cell in cells)


def _is_total_row(cost_code: str, category: str, cost_item: str) -> bool:
    text = f"{cost_code} {category} {cost_item}".lower()
    return "subtotal" in text or "grand total" in text or text.strip() in {"total", "total ex gst"}


def _normalise_category(value: str) -> str:
    lower = value.lower()
    if "pc" in lower or "prime cost" in lower:
        return "PC allowances"
    if "fee" in lower or "charge" in lower:
        return "Fees and charges"
    if "consult" in lower:
        return "Consultants"
    if "conting" in lower or "allowance" in lower:
        return "Contingency / allowances"
    if "construct" in lower or "works" in lower or "trade" in lower:
        return "Construction"
    return value or "Construction"


def _group_cost_items(items: list[CostPlanLine]) -> list[tuple[str, list[CostPlanLine]]]:
    grouped: dict[str, list[CostPlanLine]] = defaultdict(list)
    for item in items:
        grouped[item.category].append(item)

    ordered = [(category, grouped.pop(category, [])) for category in CATEGORY_ORDER]
    ordered.extend((category, grouped[category]) for category in sorted(grouped))
    return ordered


def _parse_money(value: str) -> float | None:
    cleaned = _clean_markdown_cell(value)
    lower = cleaned.lower()
    if not cleaned or "tbc" in lower or lower in {"-", "--", "n/a", "na"}:
        return None
    match = re.search(r"\(?\s*-?\$?\s*(\d[\d,]*(?:\.\d+)?)\s*([mk])?\s*\)?", lower)
    if not match:
        return None
    amount = float(match.group(1).replace(",", ""))
    suffix = match.group(2)
    if suffix == "m":
        amount *= 1_000_000
    elif suffix == "k":
        amount *= 1_000
    if lower.strip().startswith("(") or "-" in lower[:3]:
        amount *= -1
    return amount


def _status_marks_approved(status: str) -> bool:
    lower = status.lower()
    return any(word in lower for word in ("approved", "locked", "contracted", "engaged"))


def _month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def _add_months(value: date, offset: int) -> date:
    month_index = value.month - 1 + offset
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def _coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, str) and not value.startswith("="):
        return _parse_money(value)
    return None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    return str(value).strip()


def _is_summary_total_cells(cells: tuple[Cell, ...]) -> bool:
    text = " ".join(_cell_text(cell.value).lower() for cell in cells[:3])
    return "subtotal" in text or "grand total" in text


def _argb_to_hex(value: str | None) -> str | None:
    if not value or len(value) not in {6, 8}:
        return None
    if len(value) == 8:
        value = value[2:]
    if value.upper() == "000000":
        return None
    return f"#{value.upper()}"
