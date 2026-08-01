from __future__ import annotations

import uuid
from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.cost_plan.schemas import CostPlanState, DependencySnapshot
from app.sitewise.cost_plan_workbook import (
    build_cost_plan_workbook,
    build_cost_plan_workbook_for_export,
    workbook_preview_from_bytes,
)
from tests.workflows.test_create_cost_plan import _valid_cost_plan_markdown


def test_build_cost_plan_workbook_preserves_sitewise_excel_contract() -> None:
    workbook = build_cost_plan_workbook(
        project_title="Greenfield Demo",
        markdown=_valid_cost_plan_markdown(),
        version=1,
    )

    loaded = load_workbook(BytesIO(workbook.content), data_only=False)
    assert loaded.sheetnames == ["Summary", "Invoices", "Variations"]
    assert workbook.filename == "Cost_Plan_v01.draft.xlsx"
    assert workbook.row_count == 8
    assert workbook.cost_item_lookup_count == 8

    summary = loaded["Summary"]
    assert summary["A1"].value == "Project Cost Plan - Greenfield Demo"
    assert [summary.cell(row=4, column=index).value for index in range(1, 13)] == [
        "Cost Code",
        "Category",
        "Cost Items",
        "Budget",
        "Approved Contract",
        "Forecast Variations",
        "Approved Variations",
        "Forecast Final Cost",
        "Budget Variance",
        "Claimed to Date",
        "This Month",
        "Remaining",
    ]
    assert summary["F5"].value == (
        "=SUMIFS(Variations!$E$5:$E$500,Variations!$B$5:$B$500,C5,"
        'Variations!$D$5:$D$500,"<>Approved")'
    )
    assert (
        summary["G5"].value
        == "=SUMIF(Variations!$B$5:$B$500,C5,Variations!$G$5:$G$500)"
    )
    assert summary["H5"].value == "=SUM(E5:G5)"
    assert summary["K5"].value == (
        "=SUMIFS(Invoices!$G$5:$G$500,Invoices!$F$5:$F$500,C5,"
        'Invoices!$H$5:$H$500,">="&EOMONTH(K$2,-1)+1,'
        'Invoices!$H$5:$H$500,"<="&EOMONTH(K$2,0))'
    )
    assert summary.column_dimensions["M"].hidden is True
    assert summary.column_dimensions["N"].hidden is True
    assert "CostItemLookup" in loaded.defined_names
    assert "InvoiceBillingMonths" in loaded.defined_names
    assert _has_validation(summary, "K2", "InvoiceBillingMonths")
    assert _has_validation(loaded["Invoices"], "F5:F500", "CostItemLookup")
    assert _has_validation(loaded["Variations"], "B5:B500", "CostItemLookup")


def test_workbook_preview_rolls_up_invoice_and_variation_values() -> None:
    workbook = build_cost_plan_workbook(
        project_title="Greenfield Demo",
        markdown=_valid_cost_plan_markdown(),
        version=1,
    )
    loaded = load_workbook(BytesIO(workbook.content), data_only=False)
    summary = loaded["Summary"]
    invoices = loaded["Invoices"]
    variations = loaded["Variations"]
    summary["K2"] = summary["N14"].value

    invoices["F5"] = "Planning fees"
    invoices["G5"] = 3000
    invoices["H5"] = summary["K2"].value
    variations["B5"] = "Planning fees"
    variations["D5"] = "Forecast"
    variations["E5"] = 1500

    buffer = BytesIO()
    loaded.save(buffer)
    preview = workbook_preview_from_bytes(buffer.getvalue())

    summary_preview = next(sheet for sheet in preview.sheets if sheet.name == "Summary")
    planning_row = next(
        row
        for row in summary_preview.rows
        if row[:3] == ["1", "Fees and charges", "Planning fees"]
    )
    assert planning_row[5] == "$1,500"
    assert planning_row[8] == "$10,500"
    assert planning_row[9] == "$3,000"
    assert planning_row[10] == "$3,000"
    assert planning_row[11] == "$9,000"


def test_export_keeps_greenbank_tbc_rows_when_typed_import_is_empty() -> None:
    markdown = _greenbank_cost_breakdown_markdown()
    typed_state = CostPlanState(
        project_id=uuid.uuid4(),
        version=1,
        dependency_snapshot=DependencySnapshot(
            profile_revision=1,
            evidence_fingerprint="greenbank-test",
            decision_set_revision=1,
            runtime_version="test",
        ),
        items=[],
    )

    workbook = build_cost_plan_workbook_for_export(
        project_title="Greenbank",
        markdown=markdown,
        version=1,
        typed_state=typed_state,
        generated_at=datetime(2026, 8, 1, tzinfo=UTC),
    )

    loaded = load_workbook(BytesIO(workbook.content), data_only=False)
    summary = loaded["Summary"]
    item_rows = [
        (summary.cell(row=row, column=2).value, summary.cell(row=row, column=3).value)
        for row in range(5, summary.max_row + 1)
        if summary.cell(row=row, column=1).value is not None
    ]

    assert workbook.row_count == 25
    assert loaded.sheetnames == ["Summary", "Invoices", "Variations"]
    assert item_rows == _greenbank_cost_items()


def _has_validation(sheet, cell_range: str, formula: str) -> bool:
    return any(
        cell_range in str(validation.sqref) and validation.formula1 == formula
        for validation in sheet.data_validations.dataValidation
    )


def _greenbank_cost_breakdown_markdown() -> str:
    rows = "\n".join(
        f"| {index} | {category} | {item} | TBC | TBC | Assumption |"
        for index, (category, item) in enumerate(_greenbank_cost_items(), start=1)
    )
    return f"""# Project Cost Plan

## Budget reconciliation and cost breakdown

### Cost breakdown

| Cost Code | Category | Cost Items | Budget | Approved Contract | Basis |
|---|---|---|---:|---:|---|
{rows}
"""


def _greenbank_cost_items() -> list[tuple[str, str]]:
    return [
        ("Fees and charges", "Architect-PM architect / PM fee"),
        ("Fees and charges", "DA and CC authority fees"),
        ("Fees and charges", "BASIX certificate fee"),
        ("Fees and charges", "Sydney Water / infrastructure"),
        ("Fees and charges", "Levies and statutory"),
        ("Consultants", "Structural engineer"),
        ("Consultants", "Geotechnical engineer"),
        ("Consultants", "Surveyor"),
        ("Consultants", "Hydraulic / wastewater"),
        ("Consultants", "BASIX / energy assessor"),
        ("Consultants", "Principal certifier"),
        ("Construction", "Investigations, surveys and opening-up"),
        ("Construction", "Preliminaries, protection and temporary works"),
        ("Construction", "Hazardous-material controls and demolition"),
        ("Construction", "Existing-structure repair and new structural work"),
        ("Construction", "Envelope, roofing and old-to-new weatherproofing"),
        ("Construction", "Partitions, linings, doors and joinery"),
        ("Construction", "Kitchen, bathrooms and fittings"),
        ("Construction", "Building-services alterations and upgrades"),
        ("Construction", "Finishes, external works and making good"),
        ("PC allowances", "Kitchen joinery PC"),
        ("PC allowances", "Wet area / sanitary PC"),
        ("PC allowances", "Floor coverings PC"),
        ("PC allowances", "Lighting fittings PC"),
        ("Contingency / allowances", "Owner-held contingency"),
    ]
